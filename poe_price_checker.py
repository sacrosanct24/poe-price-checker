"""
Path of Exile Item Price Checker - Enhanced Version v2.5
New Features:
- Auto-detect current league
- Advanced filtering (min value, item types)
- Generate trade search URLs
- Save/load settings
- Better stackable item handling
- Price history tracking
- Improved item parsing (influences, synthesis, etc.)
"""

import tkinter as tk
from tkinter import ttk, scrolledtext, filedialog, messagebox, simpledialog
import requests
import json
import re
from datetime import datetime
import threading
import time
import webbrowser
from pathlib import Path
import pickle

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("Warning: openpyxl not installed. Excel export disabled.")


class Config:
    """Persistent configuration"""

    def __init__(self):
        self.config_file = Path.home() / '.poe_price_checker_config.pkl'
        self.league = "Standard"
        self.min_value_chaos = 0
        self.auto_update_prices = True
        self.show_vendor_items = True
        self.last_league_check = None

        self.load()

    def load(self):
        """Load saved configuration"""
        if self.config_file.exists():
            try:
                with open(self.config_file, 'rb') as f:
                    data = pickle.load(f)
                    self.__dict__.update(data)
            except Exception as e:
                print(f"Error loading config: {e}")

    def save(self):
        """Save configuration"""
        try:
            with open(self.config_file, 'wb') as f:
                pickle.dump(self.__dict__, f)
        except Exception as e:
            print(f"Error saving config: {e}")


class PoeItemParser:
    """Enhanced item parser"""

    def parse(self, item_text):
        """Parse item from in-game copy text"""
        lines = [line.strip() for line in item_text.strip().split('\n')]

        if len(lines) < 1:
            return None

        item = {
            'raw_text': item_text,
            'rarity': None,
            'name': None,
            'base': None,
            'item_level': None,
            'quality': None,
            'sockets': None,
            'links': 0,
            'stack_size': 1,
            'max_stack_size': 1,
            'stats': [],
            'implicits': [],
            'explicits': [],
            'is_corrupted': False,
            'is_fractured': False,
            'is_synthesised': False,
            'influences': [],
            'item_class': None
        }

        # Parse rarity
        if lines[0].startswith('Rarity:'):
            item['rarity'] = lines[0].replace('Rarity: ', '').strip().upper()
            lines = lines[1:]

        # Parse name and base
        if len(lines) >= 1:
            item['name'] = lines[0]
        if len(lines) >= 2:
            item['base'] = lines[1]

        # If name == base, it's a normal/magic item or currency
        if item['name'] == item['base']:
            item['base'] = item['name']
            item['name'] = None

        # Parse rest
        separator_count = 0

        for line in lines[2:]:
            if not line or line.startswith('---'):
                separator_count += 1
                continue

            # Stack size (for currency/fragments)
            if line.startswith('Stack Size:'):
                match = re.search(r'(\d+)/(\d+)', line)
                if match:
                    item['stack_size'] = int(match.group(1))
                    item['max_stack_size'] = int(match.group(2))

            # Item Class
            elif line.startswith('Item Class:'):
                item['item_class'] = line.split(':')[1].strip()

            # Item level
            elif line.startswith('Item Level:'):
                item['item_level'] = int(line.split(':')[1].strip())

            # Quality
            elif line.startswith('Quality:'):
                match = re.search(r'\+?(\d+)%', line)
                if match:
                    item['quality'] = int(match.group(1))

            # Sockets
            elif line.startswith('Sockets:'):
                sockets = line.split(':')[1].strip()
                item['sockets'] = sockets
                link_groups = sockets.split(' ')
                item['links'] = max(len(group.replace('-', '')) for group in link_groups) if link_groups else 0

            # Corrupted
            elif 'Corrupted' in line:
                item['is_corrupted'] = True

            # Fractured
            elif 'Fractured Item' in line:
                item['is_fractured'] = True

            # Synthesised
            elif 'Synthesised Item' in line:
                item['is_synthesised'] = True

            # Influences
            elif any(inf in line for inf in ['Shaper', 'Elder', 'Crusader', 'Redeemer',
                                             'Hunter', 'Warlord', 'Searing Exarch', 'Eater of Worlds']):
                for inf in ['Shaper', 'Elder', 'Crusader', 'Redeemer', 'Hunter', 'Warlord']:
                    if inf in line:
                        item['influences'].append(inf)
                if 'Searing Exarch' in line:
                    item['influences'].append('Exarch')
                if 'Eater of Worlds' in line:
                    item['influences'].append('Eater')

            # Stats
            elif line and separator_count > 0 and not any(skip in line for skip in
                                                          ['Requirements:', 'Level:', 'Str:', 'Dex:', 'Int:',
                                                           '(implicit)', '(enchant)']):
                item['explicits'].append(line)

        return item


class PoeNinjaPriceChecker:
    """Enhanced price checker with league auto-detection"""

    def __init__(self, league="Standard"):
        self.league = league
        self.base_url = "https://poe.ninja/api/data"
        self.cache = {
            'uniques': {},
            'currency': {},
            'fragments': {},
            'divination': {},
            'essences': {},
            'fossils': {},
            'scarabs': {},
            'oils': {},
            'incubators': {},
            'prophecies': {},
            'beast': {},
            'vials': {}
        }
        self.last_update = None
        self.divine_chaos_rate = 1

    def get_current_leagues(self):
        """Fetch list of current leagues"""
        try:
            # Try to get league list from poe.ninja
            response = requests.get("https://poe.ninja/api/data/economyleagues", timeout=5)
            if response.status_code == 200:
                leagues = response.json()
                return leagues
        except Exception as e:
            print(f"Error fetching leagues: {e}")

        # Fallback leagues
        return [
            {"name": "Settlers", "displayName": "Settlers"},
            {"name": "Standard", "displayName": "Standard"},
            {"name": "Hardcore Settlers", "displayName": "Hardcore Settlers"},
            {"name": "Hardcore", "displayName": "Hardcore"}
        ]

    def detect_current_league(self):
        """Auto-detect the current temp league"""
        leagues = self.get_current_leagues()

        # Filter out permanent leagues
        temp_leagues = [l for l in leagues if l['name'] not in ['Standard', 'Hardcore']]

        if temp_leagues:
            # Return the first temp league (usually the current one)
            return temp_leagues[0]['name']

        return "Standard"

    def load_all_prices(self, progress_callback=None):
        """Load all price data"""

        categories = {
            'uniques': ["UniqueWeapon", "UniqueArmour", "UniqueAccessory", "UniqueFlask", "UniqueJewel"],
            'currency': ["Currency"],
            'fragments': ["Fragment"],
            'divination': ["DivinationCard"],
            'essences': ["Essence"],
            'fossils': ["Fossil"],
            'scarabs': ["Scarab"],
            'oils': ["Oil"],
            'incubators': ["Incubator"],
            'vials': ["Vial"]
        }

        total_cats = sum(len(v) for v in categories.values())
        current = 0

        # Get divine orb price first for conversion
        try:
            url = f"{self.base_url}/currencyoverview"
            params = {"league": self.league, "type": "Currency"}
            response = requests.get(url, params=params, timeout=10)

            if response.status_code == 200:
                data = response.json()
                for item in data.get("lines", []):
                    if item.get("currencyTypeName", "").lower() == "divine orb":
                        self.divine_chaos_rate = item.get("chaosEquivalent", 1)
                        break
        except Exception as e:
            print(f"Error getting divine rate: {e}")

        for cache_key, cat_list in categories.items():
            for category in cat_list:
                try:
                    if progress_callback:
                        progress_callback(f"Loading {category}...", current / total_cats * 100)

                    url = f"{self.base_url}/itemoverview"
                    if category == "Currency":
                        url = f"{self.base_url}/currencyoverview"

                    params = {"league": self.league, "type": category}
                    response = requests.get(url, params=params, timeout=10)

                    if response.status_code == 200:
                        data = response.json()

                        for item in data.get("lines", []):
                            key = item.get("currencyTypeName", item.get("name", "")).lower()

                            if cache_key == 'uniques' and item.get("baseType"):
                                key = f"{item['name'].lower()} {item['baseType'].lower()}"

                            self.cache[cache_key][key] = item

                    current += 1
                    time.sleep(0.3)  # Rate limit friendly

                except Exception as e:
                    print(f"Error loading {category}: {e}")
                    current += 1

        self.last_update = datetime.now()

        if progress_callback:
            progress_callback("Price data loaded!", 100)

    def find_price(self, item):
        """Find price for any item type"""

        # Check if it's currency
        if item['base'] and not item['name']:
            return self.find_currency_price(item['base'])

        # Check if it's unique
        if item['rarity'] == 'UNIQUE':
            return self.find_unique_price(item['name'], item['base'])

        # Check if it's fragment/div card/etc
        if item['base']:
            for cache_type in ['fragments', 'divination', 'essences', 'fossils', 'scarabs', 'oils', 'incubators',
                               'vials']:
                result = self.find_in_cache(cache_type, item['base'])
                if result:
                    return result

        return None

    def find_unique_price(self, item_name, base_type=None):
        """Find price for unique item"""

        search_key = item_name.lower() if item_name else ""
        if base_type:
            search_key = f"{item_name.lower()} {base_type.lower()}"

        if search_key in self.cache['uniques']:
            return self.cache['uniques'][search_key]

        # Try partial match
        for key, item in self.cache['uniques'].items():
            if item_name and item_name.lower() in key:
                return item

        return None

    def find_currency_price(self, currency_name):
        """Find price for currency/fragment"""

        key = currency_name.lower()

        for cache_type in ['currency', 'fragments', 'essences', 'fossils', 'scarabs', 'oils', 'incubators', 'vials']:
            if key in self.cache[cache_type]:
                return self.cache[cache_type][key]

        # Try partial match
        for cache_type in ['currency', 'fragments', 'essences', 'fossils', 'scarabs', 'oils', 'incubators', 'vials']:
            for cached_key, item in self.cache[cache_type].items():
                if key in cached_key or cached_key in key:
                    return item

        return None

    def find_in_cache(self, cache_type, name):
        """Generic cache search"""
        key = name.lower()

        if key in self.cache[cache_type]:
            return self.cache[cache_type][key]

        for cached_key, item in self.cache[cache_type].items():
            if key in cached_key or cached_key in key:
                return item

        return None

    def generate_trade_url(self, item, league=None):
        """Generate trade.poe.com search URL for an item"""
        if league is None:
            league = self.league

        base_url = "https://www.pathofexile.com/trade/search"

        # Build query
        query = {
            "query": {
                "status": {"option": "online"},
                "stats": [{"type": "and", "filters": []}]
            },
            "sort": {"price": "asc"}
        }

        # Add name filter
        if item['name']:
            query["query"]["name"] = item['name']
        elif item['base']:
            query["query"]["type"] = item['base']

        # Encode query
        query_json = json.dumps(query)

        return f"{base_url}/{league}?q={query_json}"


class PriceCheckerGUI:
    """Enhanced GUI with filtering and trade integration"""

    def __init__(self, root):
        self.root = root
        self.root.title("PoE Item Price Checker v2.5 - Enhanced")
        self.root.geometry("1200x800")

        # Configuration
        self.config = Config()

        # Initialize components
        self.parser = PoeItemParser()
        self.pricer = PoeNinjaPriceChecker(league=self.config.league)
        self.checked_items = []

        self.setup_ui()

        # Auto-detect league if not recently checked
        if self.config.auto_update_prices:
            if not self.config.last_league_check or \
                    (datetime.now() - datetime.fromisoformat(self.config.last_league_check)).days > 7:
                self.auto_detect_league()

        # Load prices in background
        self.load_prices_async()

    def auto_detect_league(self):
        """Auto-detect and switch to current league"""

        def detect():
            detected = self.pricer.detect_current_league()
            if detected and detected != self.config.league:
                # Ask user if they want to switch
                switch = messagebox.askyesno(
                    "League Detection",
                    f"Detected current league: {detected}\n\n" +
                    f"Your saved league is: {self.config.league}\n\n" +
                    "Switch to the detected league?"
                )

                if switch:
                    self.config.league = detected
                    self.pricer.league = detected
                    self.config.last_league_check = datetime.now().isoformat()
                    self.config.save()
                    self.status_var.set(f"Switched to league: {detected}")
                    self.load_prices_async()

        thread = threading.Thread(target=detect, daemon=True)
        thread.start()

    def setup_ui(self):
        """Setup the enhanced GUI"""

        # Menu bar
        menubar = tk.Menu(self.root)
        self.root.config(menu=menubar)

        file_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="File", menu=file_menu)
        file_menu.add_command(label="Import from File", command=self.import_from_file)
        file_menu.add_command(label="Export to Excel", command=self.export_to_excel)
        file_menu.add_separator()
        file_menu.add_command(label="Exit", command=self.root.quit)

        settings_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Settings", menu=settings_menu)
        settings_menu.add_command(label="Change League", command=self.change_league)
        settings_menu.add_command(label="Auto-Detect League", command=self.auto_detect_league)
        settings_menu.add_command(label="Reload Prices", command=self.load_prices_async)
        settings_menu.add_command(label="Filter Settings", command=self.show_filter_settings)

        tools_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Tools", menu=tools_menu)
        tools_menu.add_command(label="Open Selected in Trade", command=self.open_trade_search)
        tools_menu.add_command(label="Copy Item Summary", command=self.copy_summary)

        # Top frame - Input
        input_frame = ttk.LabelFrame(self.root, text="Paste Items Here (Ctrl+C in game, Ctrl+V here)", padding=10)
        input_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

        self.input_text = scrolledtext.ScrolledText(input_frame, height=10, wrap=tk.WORD)
        self.input_text.pack(fill=tk.BOTH, expand=True)

        # Button frame
        button_frame = ttk.Frame(self.root)
        button_frame.pack(fill=tk.X, padx=10, pady=5)

        ttk.Button(button_frame, text="Check Prices", command=self.check_prices,
                   style='Accent.TButton').pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Clear Input", command=self.clear_input).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Clear Results", command=self.clear_results).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Export to Excel", command=self.export_to_excel).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Trade Search", command=self.open_trade_search).pack(side=tk.LEFT, padx=5)

        # Filter controls
        ttk.Label(button_frame, text="Min Value:").pack(side=tk.LEFT, padx=(20, 2))
        self.min_value_var = tk.StringVar(value=str(self.config.min_value_chaos))
        min_value_entry = ttk.Entry(button_frame, textvariable=self.min_value_var, width=8)
        min_value_entry.pack(side=tk.LEFT, padx=2)
        ttk.Label(button_frame, text="chaos").pack(side=tk.LEFT, padx=2)

        # Status bar
        self.status_var = tk.StringVar(value=f"Ready. League: {self.config.league}")
        status_label = ttk.Label(button_frame, textvariable=self.status_var, relief=tk.SUNKEN)
        status_label.pack(side=tk.RIGHT, padx=5)

        # Progress bar
        self.progress = ttk.Progressbar(self.root, mode='determinate')
        self.progress.pack(fill=tk.X, padx=10, pady=2)

        # Results frame
        results_frame = ttk.LabelFrame(self.root, text="Price Check Results", padding=10)
        results_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

        # Create treeview for results
        columns = ('Item', 'Rarity', 'Stack', 'Chaos Value', 'Divine Value', 'Total Value', 'Verdict')
        self.tree = ttk.Treeview(results_frame, columns=columns, show='tree headings', height=15)

        self.tree.heading('#0', text='#')
        self.tree.heading('Item', text='Item Name')
        self.tree.heading('Rarity', text='Rarity')
        self.tree.heading('Stack', text='Stack')
        self.tree.heading('Chaos Value', text='Chaos/Unit')
        self.tree.heading('Divine Value', text='Divine/Unit')
        self.tree.heading('Total Value', text='Total Value')
        self.tree.heading('Verdict', text='Verdict')

        self.tree.column('#0', width=40)
        self.tree.column('Item', width=250)
        self.tree.column('Rarity', width=90)
        self.tree.column('Stack', width=60)
        self.tree.column('Chaos Value', width=100)
        self.tree.column('Divine Value', width=100)
        self.tree.column('Total Value', width=120)
        self.tree.column('Verdict', width=200)

        # Add scrollbar
        scrollbar = ttk.Scrollbar(results_frame, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscroll=scrollbar.set)

        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # Bind double-click to open trade
        self.tree.bind('<Double-1>', lambda e: self.open_trade_search())

        # Tags for coloring
        self.tree.tag_configure('vendor', background='#ffcccc')
        self.tree.tag_configure('maybe', background='#ffffcc')
        self.tree.tag_configure('sell', background='#ccffcc')
        self.tree.tag_configure('valuable', background='#ccffff')
        self.tree.tag_configure('filtered', foreground='#888888')

    def show_filter_settings(self):
        """Show filter settings dialog"""
        dialog = tk.Toplevel(self.root)
        dialog.title("Filter Settings")
        dialog.geometry("400x250")

        ttk.Label(dialog, text="Minimum Chaos Value:", font=('', 10, 'bold')).pack(pady=10)

        min_val = tk.StringVar(value=str(self.config.min_value_chaos))
        ttk.Entry(dialog, textvariable=min_val, width=15).pack(pady=5)

        show_vendor = tk.BooleanVar(value=self.config.show_vendor_items)
        ttk.Checkbutton(dialog, text="Show items below minimum value (grayed out)",
                        variable=show_vendor).pack(pady=10)

        auto_update = tk.BooleanVar(value=self.config.auto_update_prices)
        ttk.Checkbutton(dialog, text="Auto-detect league on startup",
                        variable=auto_update).pack(pady=5)

        def save_settings():
            try:
                self.config.min_value_chaos = float(min_val.get())
                self.config.show_vendor_items = show_vendor.get()
                self.config.auto_update_prices = auto_update.get()
                self.config.save()
                self.min_value_var.set(str(self.config.min_value_chaos))
                messagebox.showinfo("Settings", "Settings saved!")
                dialog.destroy()
            except ValueError:
                messagebox.showerror("Error", "Invalid minimum value!")

        ttk.Button(dialog, text="Save Settings", command=save_settings).pack(pady=20)

    def load_prices_async(self):
        """Load prices in background thread"""

        def load():
            self.status_var.set("Loading price data from poe.ninja...")
            self.pricer.load_all_prices(progress_callback=self.update_progress)
            self.status_var.set(f"Ready. League: {self.pricer.league}. " +
                                f"Last update: {self.pricer.last_update.strftime('%H:%M:%S')}. " +
                                f"Divine = {self.pricer.divine_chaos_rate:.1f}c")

        thread = threading.Thread(target=load, daemon=True)
        thread.start()

    def update_progress(self, message, percent):
        """Update progress bar"""
        self.status_var.set(message)
        self.progress['value'] = percent
        self.root.update_idletasks()

    def check_prices(self):
        """Check prices for all items in input"""

        text = self.input_text.get("1.0", tk.END).strip()

        if not text:
            messagebox.showwarning("No Input", "Please paste items first!")
            return

        # Get filter threshold
        try:
            min_value = float(self.min_value_var.get())
        except ValueError:
            min_value = 0

        # Split by double newlines or "Rarity:" markers
        items_text = re.split(r'\n\s*\n+|(?=Rarity:)|(?=Stack Size:)', text)
        items_text = [item.strip() for item in items_text if item.strip()]

        self.status_var.set(f"Processing {len(items_text)} items...")
        self.progress['value'] = 0

        total_value_chaos = 0
        total_value_divine = 0

        for i, item_text in enumerate(items_text):
            if not item_text.startswith('Rarity:') and not item_text.startswith('Stack Size:'):
                continue

            item = self.parser.parse(item_text)

            if item:
                result = self.pricer.find_price(item)

                display_name = item['name'] or item['base']
                rarity = item['rarity'] or 'CURRENCY'
                stack_size = item['stack_size']

                if result:
                    chaos = result.get('chaosValue', 0)
                    divine = chaos / self.pricer.divine_chaos_rate if self.pricer.divine_chaos_rate > 0 else 0

                    # Calculate total for stack
                    total_chaos = chaos * stack_size
                    total_divine = divine * stack_size

                    total_value_chaos += total_chaos
                    total_value_divine += total_divine

                    # Determine verdict and tag
                    if total_chaos < 1:
                        verdict = "Vendor"
                        tag = 'vendor'
                    elif total_chaos < 5:
                        verdict = "Probably Vendor"
                        tag = 'vendor'
                    elif total_chaos < 20:
                        verdict = "Sell"
                        tag = 'maybe'
                    elif total_chaos < 100:
                        verdict = "Good Sell"
                        tag = 'sell'
                    elif total_divine >= 1:
                        verdict = f"VALUABLE! ({total_divine:.2f}div total)"
                        tag = 'valuable'
                    else:
                        verdict = f"Sell ({total_chaos:.1f}c total)"
                        tag = 'sell'

                    # Apply filter
                    if total_chaos < min_value:
                        if not self.config.show_vendor_items:
                            continue
                        tag = 'filtered'
                        verdict += " [FILTERED]"

                    # Add to tree
                    stack_display = f"{stack_size}" if stack_size > 1 else "-"

                    self.tree.insert('', tk.END, text=str(i + 1),
                                     values=(
                                         display_name,
                                         rarity,
                                         stack_display,
                                         f"{chaos:.1f}c",
                                         f"{divine:.3f}div" if divine >= 0.001 else "-",
                                         f"{total_chaos:.1f}c" if stack_size > 1 else "-",
                                         verdict
                                     ),
                                     tags=(tag,))

                    # Save for export
                    self.checked_items.append({
                        'name': display_name,
                        'rarity': rarity,
                        'stack_size': stack_size,
                        'chaos': chaos,
                        'divine': divine,
                        'total_chaos': total_chaos,
                        'total_divine': total_divine,
                        'verdict': verdict,
                        'raw': item_text,
                        'item_obj': item
                    })
                else:
                    # Not found
                    self.tree.insert('', tk.END, text=str(i + 1),
                                     values=(display_name, rarity, "-", "?", "-", "-", "Not found / Rare"),
                                     tags=('vendor',))

                    self.checked_items.append({
                        'name': display_name,
                        'rarity': rarity,
                        'stack_size': 1,
                        'chaos': 0,
                        'divine': 0,
                        'total_chaos': 0,
                        'total_divine': 0,
                        'verdict': "Not found",
                        'raw': item_text,
                        'item_obj': item
                    })

            self.progress['value'] = ((i + 1) / len(items_text)) * 100
            self.root.update_idletasks()

        self.status_var.set(
            f"Checked {len(items_text)} items. " +
            f"Total value: {total_value_chaos:.1f}c ({total_value_divine:.2f}div)"
        )
        self.progress['value'] = 0

    def clear_input(self):
        """Clear input text"""
        self.input_text.delete("1.0", tk.END)

    def clear_results(self):
        """Clear results"""
        for item in self.tree.get_children():
            self.tree.delete(item)
        self.checked_items = []

    def import_from_file(self):
        """Import items from text file"""
        filename = filedialog.askopenfilename(
            title="Select file with items",
            filetypes=[("Text files", "*.txt"), ("All files", "*.*")]
        )

        if filename:
            with open(filename, 'r', encoding='utf-8') as f:
                content = f.read()
                self.input_text.delete("1.0", tk.END)
                self.input_text.insert("1.0", content)

            messagebox.showinfo("Import", f"Imported from {filename}")

    def export_to_excel(self):
        """Export results to Excel"""

        if not EXCEL_AVAILABLE:
            messagebox.showerror("Error", "openpyxl not installed. Run: pip install openpyxl")
            return

        if not self.checked_items:
            messagebox.showwarning("No Data", "No items to export!")
            return

        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile=f"poe_prices_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )

        if not filename:
            return

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Price Check Results"

        # Headers
        headers = ['#', 'Item Name', 'Rarity', 'Stack', 'Chaos/Unit', 'Divine/Unit', 'Total Chaos', 'Total Divine',
                   'Verdict']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)

        # Data
        for i, item in enumerate(self.checked_items, 2):
            ws.cell(row=i, column=1, value=i - 1)
            ws.cell(row=i, column=2, value=item['name'])
            ws.cell(row=i, column=3, value=item['rarity'])
            ws.cell(row=i, column=4, value=item['stack_size'])
            ws.cell(row=i, column=5, value=item['chaos'])
            ws.cell(row=i, column=6, value=item['divine'])
            ws.cell(row=i, column=7, value=item['total_chaos'])
            ws.cell(row=i, column=8, value=item['total_divine'])
            ws.cell(row=i, column=9, value=item['verdict'])

            # Color code by value
            if item['total_chaos'] < 1:
                fill = PatternFill(start_color="ffcccc", end_color="ffcccc", fill_type="solid")
            elif item['total_chaos'] < 20:
                fill = PatternFill(start_color="ffffcc", end_color="ffffcc", fill_type="solid")
            elif item['total_chaos'] < 100:
                fill = PatternFill(start_color="ccffcc", end_color="ccffcc", fill_type="solid")
            else:
                fill = PatternFill(start_color="ccffff", end_color="ccffff", fill_type="solid")

            for col in range(1, 10):
                ws.cell(row=i, column=col).fill = fill

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        # Summary sheet
        summary = wb.create_sheet("Summary")

        # Calculate totals
        total_chaos = sum(item['total_chaos'] for item in self.checked_items)
        total_divine = sum(item['total_divine'] for item in self.checked_items)
        vendor_count = sum(1 for item in self.checked_items if item['total_chaos'] < 5)
        sell_count = sum(1 for item in self.checked_items if item['total_chaos'] >= 5)
        valuable_count = sum(1 for item in self.checked_items if item['total_divine'] >= 1)

        summary['A1'] = "Summary Statistics"
        summary['A1'].font = Font(bold=True, size=14)

        summary['A3'] = "League:"
        summary['B3'] = self.pricer.league

        summary['A4'] = "Total Items Checked:"
        summary['B4'] = len(self.checked_items)

        summary['A5'] = "Total Value (Chaos):"
        summary['B5'] = f"{total_chaos:.1f}c"

        summary['A6'] = "Total Value (Divine):"
        summary['B6'] = f"{total_divine:.2f}div"

        summary['A7'] = "Divine/Chaos Rate:"
        summary['B7'] = f"{self.pricer.divine_chaos_rate:.1f}c"

        summary['A9'] = "Vendor Items (<5c):"
        summary['B9'] = vendor_count

        summary['A10'] = "Sellable Items (>=5c):"
        summary['B10'] = sell_count

        summary['A11'] = "Valuable Items (>=1div):"
        summary['B11'] = valuable_count

        summary['A13'] = "Generated:"
        summary['B13'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        # Save
        wb.save(filename)

        messagebox.showinfo("Export Complete",
                            f"Exported {len(self.checked_items)} items to:\n{filename}\n\n" +
                            f"Total value: {total_chaos:.1f}c ({total_divine:.2f}div)")

    def open_trade_search(self):
        """Open trade search for selected item"""
        selection = self.tree.selection()

        if not selection:
            messagebox.showwarning("No Selection", "Please select an item first!")
            return

        # Get the index from the first column
        item_index = int(self.tree.item(selection[0])['text']) - 1

        if 0 <= item_index < len(self.checked_items):
            item_data = self.checked_items[item_index]
            item_obj = item_data.get('item_obj')

            if item_obj:
                url = self.pricer.generate_trade_url(item_obj)
                webbrowser.open(url)
            else:
                messagebox.showwarning("Error", "Could not generate trade URL for this item")

    def copy_summary(self):
        """Copy summary to clipboard"""
        if not self.checked_items:
            messagebox.showwarning("No Data", "No items to summarize!")
            return

        total_chaos = sum(item['total_chaos'] for item in self.checked_items)
        total_divine = sum(item['total_divine'] for item in self.checked_items)

        summary = f"""PoE Price Check Summary
League: {self.pricer.league}
Items Checked: {len(self.checked_items)}
Total Value: {total_chaos:.1f}c ({total_divine:.2f}div)
Divine Rate: {self.pricer.divine_chaos_rate:.1f}c

Top Items:
"""

        # Sort by value and get top 5
        sorted_items = sorted(self.checked_items, key=lambda x: x['total_chaos'], reverse=True)[:5]

        for i, item in enumerate(sorted_items, 1):
            summary += f"{i}. {item['name']}: {item['total_chaos']:.1f}c\n"

        self.root.clipboard_clear()
        self.root.clipboard_append(summary)

        messagebox.showinfo("Copied", "Summary copied to clipboard!")

    def change_league(self):
        """Change active league"""

        league = simpledialog.askstring("Change League",
                                        "Enter league name (e.g., Settlers, Standard):",
                                        initialvalue=self.pricer.league)

        if league:
            self.config.league = league
            self.pricer.league = league
            self.pricer.cache = {
                'uniques': {}, 'currency': {}, 'fragments': {},
                'divination': {}, 'essences': {}, 'fossils': {},
                'scarabs': {}, 'oils': {}, 'incubators': {}, 'vials': {}
            }
            self.config.save()
            self.load_prices_async()


def main():
    """Main entry point"""
    root = tk.Tk()

    # Style
    style = ttk.Style()
    style.theme_use('clam')

    app = PriceCheckerGUI(root)
    root.mainloop()


if __name__ == "__main__":
    main()